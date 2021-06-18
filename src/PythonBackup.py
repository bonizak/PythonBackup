import argparse
import os
import sys
import tarfile
import time

from openpyxl import load_workbook

import Target_File_Builder as tfb
from FSWalker import FSWalker as fsw
from CommonOs import OsServices as os_services


class PythonBackup(os_services):
    def __init__(self):
        super().__init__()
        self.BackupSet_AoD = []
        self.StorageSet_AoD = []
        self.FileSet_AoD = []
        self.args = ""
        self.resource_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "resource")

    def parseCommandLine(self):
        """
        This method sets up, parses and returns the command line arguments passed to script, it inherits
        common command line arguments such as notify ops from the CommonArgs.py
         """
        try:
            parser = argparse.ArgumentParser(prog=str(sys.argv[0]),
                                             usage="%(prog)s Backups a computer's file systems according to the "
                                                   "information provided in a companion spreadsheet.")
            parser.add_argument("run_type",
                                help=f'Pass the backup frequency from DAILY, WEEKLY, MONTHLY, ARCHIVE or ANY')
            parser.add_argument("-reload", action="store_true", required=False,
                                help=f'Pass to reload the FileSet sheet of the BackupSetList.xlsx in the '
                                     f'resources directory. Review and edit the FileSets sheet once reloaded.')

            self.args = parser.parse_args()
        except Exception as e:
            print(
                f'An exception occurred which caused the command line argument to not parse, investigate {e}')
        else:
            return self.args

    def run_object(self):
        if args.reload:
            fsw_rc = fsw.Build_FileSets()
            if fsw_rc > 0:
                os_services.info(self, f"Reload of FileSets in BackupList.xlsx loaded {fsw} rows.")
            else:
                os_services.error(self, f"Reload of FileSets in BackupList.xlsx failed.")
        else:
            os_services.info(self, f'Backing up host {os.uname()[1]}.')
            self.BackupSet_AoD, self.StorageSet_AoD, self.FileSet_AoD = self.excel_convert()
            self.backup_start(str(args.run_type).upper())
            os_services.info(self, f'Completed {os.uname()[1]} backup')

    def backup_start(self, run_type):
        backup_list_in = self.BackupSet_AoD

        # loop through backup_sets
        for index in range(len(backup_list_in)):
            backup_set_name = ""
            file_set_name = ""
            storage_path = ""
            backup_versions = 1
            frequency = ""
            include_files_list = []
            exclude_files_list = []
            skipping = True
            for key in backup_list_in[index]:  # loop through backup set key fields
                if key == "BackupSetName":
                    backup_set_name = backup_list_in[index][key]
                    os_services.info(self, f'Running BackupSet \'{backup_set_name}\' ')
                elif key == "StorageSetName":
                    storage_path = self.storage_path_getter(backup_list_in[index][key])
                    if storage_path is None:
                        os_services.error(self, "Empty StoragePath. Skipping BackupSet")
                        break
                elif key == "FileSetName":
                    include_files_list = self.fileset_includes_getter(backup_list_in[index][key])
                    exclude_files_list = self.fileset_excludes_getter(backup_list_in[index][key])
                    file_set_name = backup_list_in[index][key]
                elif key == "Frequency" and run_type.upper() == str(backup_list_in[index][key]).upper():
                    frequency = str(backup_list_in[index][key]).upper()
                    skipping = False
                elif key == "Versions":
                    backup_versions = backup_list_in[index][key]
                else:
                    pass

            try:
                if skipping:
                    os_services.info(self,
                                     f" Skipping Backup Set \'{backup_set_name}\' as it is not scheduled for today.\n")
                    continue
                else:
                    os_services.info(self, f' FileSet: {file_set_name} ')
                    os_services.info(self, f' StoragePath: {storage_path}')
                    os_services.debug(self, f' Frequency: {frequency}')
                    os_services.debug(self, f'  Includes: {include_files_list} ')
                    os_services.debug(self, f'  Excludes: {exclude_files_list}')
            except Exception as ex:
                os_services.error(self, f"Exception with BackupSet {backup_set_name}. {ex}")
                os_services.error(self, f' FileSet: {file_set_name} ')
                os_services.error(self, f' StoragePath: {storage_path}')
                os_services.error(self, f' Frequency: {frequency}')
                os_services.error(self, f'  Includes: {include_files_list} ')
                os_services.error(self, f'  Excludes: {exclude_files_list}')
                continue

            # remove any excluded files from the includes list
            if len(exclude_files_list) > 0:
                for excl in exclude_files_list:
                    if excl in include_files_list:
                        include_files_list.remove(excl)

            # Determine the archive file name based on the current versions
            archive_target_basefile = os.path.join(storage_path, backup_set_name, f'{backup_set_name}')
            os_services.debug(self, f"Searching for the number of {archive_target_basefile}* files")
            os_services.debug(self, f"  keeping only {backup_versions} versions")
            archive_builder = tfb.Target_File_Builder(f'{archive_target_basefile}', backup_versions)
            archive_file = archive_builder.archive_target_file
            archive_rc = self.write_tar_file(archive_file, include_files_list, True)
            os_services.info(self, f'Back up of Backup Set Name {backup_set_name} '
                                   f'into {archive_file} '
                                   f'returned {archive_rc}\n')

    def storage_path_getter(self, storageSet_needle):
        storagesets_in = self.StorageSet_AoD

        for index in range(len(storagesets_in)):
            for key in storagesets_in[index]:
                if key == "StorageSetName" and storagesets_in[index]['StorageSetName'] == storageSet_needle:
                    return storagesets_in[index]["StoragePath"]

    def fileset_includes_getter(self, filesetname_needle):
        filesets_in = self.FileSet_AoD
        fs_includes = []

        for index in range(len(filesets_in)):
            for key in filesets_in[index]:
                if key == "FileSetName" and filesets_in[index]["FileSetName"] == filesetname_needle:
                    if os.path.exists(filesets_in[index]["Includes"]):
                        fs_includes.append(filesets_in[index]["Includes"])
                    else:
                        msg = f'  File Set {filesets_in[index]["Includes"]} does not exist. Remove it from ' \
                              f'FileSets sheet in BackupList.xlsx '
                        os_services.error(self, msg)
        return fs_includes

    def fileset_excludes_getter(self, filesetname_needle):
        filesets_in = self.FileSet_AoD
        fs_excludes = []

        for index in range(len(filesets_in)):
            for key in filesets_in[index]:
                if "FileSetName" in key and filesetname_needle in filesets_in[index][key]:
                    if os.path.exists(filesets_in[index]["Excludes"]):
                        fs_excludes.append(filesets_in[index]["Excludes"])

        return fs_excludes

    @staticmethod
    def is_file_older_than_x_days(file, days=1):
        file_time = os.path.getmtime(file)
        # Check against 24 hours
        return (time.time() - file_time) / 3600 > 24 * days

    @staticmethod
    def write_tar_file(target, sources, recursive):
        """ Tar and compress the sources into the target """
        try:
            with tarfile.open(target, 'w:gz') as tar_out:
                for src in sources:
                    tar_out.add(src, recursive=recursive)

            tar_out.close()
            return 0
        except OSError as oserr:
            return oserr

    def excel_convert(self):
        wb = load_workbook(os.path.join(self.resource_path, "BackupList.xlsx"))
        sheetset = {'BackupSets': 6, 'StorageSets': 4, 'FileSets': 6}

        for ws in wb:
            if ws.title in sheetset.keys():
                worksheet = wb[ws.title]

                row_sets = [worksheetsets for worksheetsets in worksheet.iter_rows(
                    min_row=2, max_col=sheetset[ws.title], min_col=1, values_only=True)
                            if None not in worksheetsets]

                row_set_count = 0
                for row_set in row_sets:
                    row_set_dict = {}
                    if worksheet.title == "BackupSets":
                        for index in range(len(row_set)):
                            if index == 0:
                                row_set_dict["Index"] = row_set_count
                            elif index == 1:
                                row_set_dict["BackupSetName"] = row_set[index]
                            elif index == 2:
                                row_set_dict["StorageSetName"] = row_set[index]
                            elif index == 3:
                                row_set_dict["FileSetName"] = row_set[index]
                            elif index == 4:
                                row_set_dict["Versions"] = row_set[index]
                            elif index == 5:
                                row_set_dict["Frequency"] = row_set[index]
                            else:
                                raise AttributeError

                        self.BackupSet_AoD.append(row_set_dict)
                        row_set_count += 1

                    elif worksheet.title == "StorageSets":
                        for index in range(len(row_set)):
                            if index == 0:
                                row_set_dict["Index"] = row_set_count
                            elif index == 1:
                                row_set_dict["StorageSetName"] = row_set[index]
                            elif index == 2:
                                row_set_dict["StoragePath"] = row_set[index]
                            elif index == 3:
                                row_set_dict["DeviceType"] = row_set[index]
                            else:
                                raise AttributeError

                        self.StorageSet_AoD.append(row_set_dict)
                        row_set_count += 1

                    elif worksheet.title == "FileSets":
                        for index in range(len(row_set)):
                            if index == 0:
                                row_set_dict["Index"] = row_set_count
                            elif index == 1:
                                row_set_dict["FileSetName"] = row_set[index]
                            elif index == 2:
                                row_set_dict["Includes"] = row_set[index]
                            elif index == 3:
                                row_set_dict["Excludes"] = row_set[index]
                            elif index == 4:
                                row_set_dict["Compress"] = row_set[index]
                            elif index == 5:
                                row_set_dict["Recurse"] = row_set[index]
                            else:
                                raise AttributeError

                        self.FileSet_AoD.append(row_set_dict)
                        row_set_count += 1

        os_services.info(self, f' read {len(self.BackupSet_AoD)} Backup sets')
        os_services.info(self, f' read {len(self.StorageSet_AoD)} Storage sets')
        os_services.info(self, f' read {len(self.FileSet_AoD)} File sets')
        return self.BackupSet_AoD, self.StorageSet_AoD, self.FileSet_AoD


# =================================
if __name__ == '__main__':
    obj = PythonBackup()
    args = obj.parseCommandLine()
    obj.getLogger(__name__)
    obj.starting_template(sys.argv[1:], args)
    obj.run_object()
    obj.ending_template(sys.argv[1:], args)
