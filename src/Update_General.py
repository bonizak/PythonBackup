import fnmatch
import os
import shutil
import sys

from openpyxl import load_workbook
from CommonOs import OsServices as os_services


class UpdateGeneral(os_services):
    """
    This class contains the methods required to copy non-user files into a user folder for backup
    """

    def Collect_General_Files(self):
        """
        This method collects the non-user files list in the GeneralList sheet
         and copies them into a $HOME/Documents/General folder
        :return:
        """
        General_AoD = self.extract_GeneralList()
        copy_count = 0

        for index in range(len(General_AoD)):
            for key in General_AoD[index]:
                if key == "SourceFile_FolderName" and "*" in General_AoD[index]["SourceFile_FolderName"]:
                    ff_dirname = os.path.dirname(General_AoD[index]["SourceFile_FolderName"])
                    ff_basename = os.path.basename(General_AoD[index]["SourceFile_FolderName"])
                    os_services.debug(self, f' Scan for {ff_basename} in Folder {ff_dirname}')
                    if os.path.exists(ff_dirname):
                        for file in os.listdir(ff_dirname):
                            if fnmatch.fnmatch(file, ff_basename):
                                os_services.debug(self, f' Copying {file} into {General_AoD[index]["TargetFolder"]}')
                                if not os.path.exists(General_AoD[index]["TargetFolder"]):
                                    os.makedirs(General_AoD[index]["TargetFolder"])
                                shutil.copy(os.path.join(ff_dirname, file), General_AoD[index]["TargetFolder"])
                                copy_count += 1
                    else:
                        os_services.critical(self, f'{ff_dirname} does not exist.')
                elif key == "SourceFile_FolderName" and os.path.isdir(General_AoD[index]["SourceFile_FolderName"]):
                    os_services.debug(self,
                                      f' row {index} is a folder {General_AoD[index]["SourceFile_FolderName"]}')
                    os_services.debug(self, f' Scan for all files in Folder '
                                            f'{General_AoD[index]["SourceFile_FolderName"]}')
                    for file in os.listdir(General_AoD[index]["SourceFile_FolderName"]):
                        os_services.debug(self, f' Copying {file} into {General_AoD[index]["TargetFolder"]}')
                        if not os.path.exists(General_AoD[index]["TargetFolder"]):
                            os.makedirs(General_AoD[index]["TargetFolder"])
                        shutil.copy2(os.path.join(General_AoD[index]["SourceFile_FolderName"], file),
                                     General_AoD[index]["TargetFolder"])
                        copy_count += 1
                elif key == "SourceFile_FolderName" and os.path.isfile(General_AoD[index]["SourceFile_FolderName"]):
                    os_services.debug(self, f' Copying {General_AoD[index]["SourceFile_FolderName"]} into '
                                            f'{General_AoD[index]["TargetFolder"]}')
                    if not os.path.exists(General_AoD[index]["TargetFolder"]):
                        os.makedirs(General_AoD[index]["TargetFolder"])
                    shutil.copy(General_AoD[index]["SourceFile_FolderName"],
                                General_AoD[index]["TargetFolder"])
                    copy_count += 1
        os_services.info(self, f' Copied {copy_count} files.')
        return copy_count

    def extract_GeneralList(self):
        resource_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "resource")
        wb = load_workbook(os.path.join(resource_path, "BackupList.xlsx"))
        sheetset = {'GeneralList': 4}

        GeneralList_AoD = []

        for ws in wb:
            if ws.title in sheetset.keys():
                worksheet = wb[ws.title]

                row_sets = [worksheetsets for worksheetsets in worksheet.iter_rows(
                    min_row=2, max_col=sheetset[ws.title], min_col=1, values_only=True)
                            if None not in worksheetsets]

                if len(row_sets) == 0:
                    print(f' No rows read from BackupList')
                    sys.exit(1)
                else:
                    row_set_count = 0
                    for row_set in row_sets:
                        row_set_dict = {}
                        if worksheet.title == "GeneralList":
                            for index in range(len(row_set)):
                                if index == 0:
                                    row_set_dict["Index"] = row_set_count
                                elif index == 1:
                                    row_set_dict["SourceFile_FolderName"] = row_set[index]
                                elif index == 2:
                                    row_set_dict["TargetFolder"] = row_set[index]
                                elif index == 3:
                                    row_set_dict["EstimatedSize"] = row_set[index]
                                else:
                                    raise AttributeError

                            GeneralList_AoD.append(row_set_dict)
                            row_set_count += 1

        os_services.debug(self, f' Read in {len(GeneralList_AoD)} File_Folders sets')
        return GeneralList_AoD
