import os

import pandas as pd
from openpyxl import load_workbook
from CommonOs import OsServices as os_services


class FSWriter(os_services):
    """
    This class contains the methods to read the FileSets sheet and
    collects and updates the 'Estimated Size' cell for each row
    
    Args
        Required: none
        Optional: none

    Logging: WARN | ERROR

    """

    __author__ = "Barry Onizak"
    __version__ = "20220330.1"
    # # # # # End of header # # # #
    
    resource_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "resource")

    def Collect_file_sizes(self):
        FileSizeRows = []
        IncludesRead = self.read_filesets_in()
        row_set_count = 0

        for index in range(len(IncludesRead)):
            for key in IncludesRead[index]:
                FileSizeRow = {}
                if key == "Includes" and os.path.isdir(IncludesRead[index][key]) \
                        and ("lost+found" or "Trash") not in IncludesRead[index][key]:
                    key_size = self.getFolderSize(IncludesRead[index][key], IncludesRead[index]["Recurse"])
                    FileSizeRow["Index"] = row_set_count
                    FileSizeRow["Path"] = IncludesRead[index][key]
                    FileSizeRow["Size"] = key_size
                    FileSizeRows.append(FileSizeRow)
                    row_set_count += 1
                elif key == "Includes" and os.path.isfile(IncludesRead[index][key]) \
                        and ("lost+found" or "Trash") not in IncludesRead[index][key]:
                    key_size = os.stat(IncludesRead[index][key]).st_size
                    FileSizeRow["Index"] = row_set_count
                    FileSizeRow["Path"] = IncludesRead[index][key]
                    FileSizeRow["Size"] = key_size
                    FileSizeRows.append(FileSizeRow)
                    row_set_count += 1
                elif key == "Includes":
                    os_services.warn(self, f' {IncludesRead[index][key]} does not exist.  '
                                     f'Consider removing it from FileSets.')
                else:
                    pass

        write_files_sizes_rc = self.write_files_sizes(FileSizeRows)
        return write_files_sizes_rc

    def getFolderSize(self, folder, recurse):
        itempath = ""
        total_size = os.path.getsize(folder)
        if str(recurse).upper() == "YES":
            try:
                for item in os.listdir(folder):
                    itempath = os.path.join(folder, item)
                    if os.path.isfile(itempath):
                        total_size += os.path.getsize(itempath)
                    elif os.path.isdir(itempath):
                        total_size += self.getFolderSize(itempath, recurse)
            except PermissionError as pe:
                os_services.error(self, f'{itempath} is not accessible. {pe}')
                return 0
            except OSError as ose:
                os_services.error(self, f'{itempath} access error. {ose}')
                return 0
        return total_size

    def read_filesets_in(self):
        """
        This method reads in the FileSets sheet and writes the specific columns into an array of dictionaries
        """
        resource_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "resource")
        wb = load_workbook(os.path.join(self.resource_path, "BackupList.xlsx"))
        sheetset = {'FileSets': 6}
        IncludesRead = []

        for ws in wb:
            if ws.title in sheetset.keys():
                worksheet = wb[ws.title]

                row_sets = [worksheetsets for worksheetsets in worksheet.iter_rows(
                    min_row=2, max_col=sheetset[ws.title], min_col=1, values_only=True)
                            if None not in worksheetsets]
                row_set_count = 0
                for row_set in row_sets:
                    row_set_dict = {}
                    if worksheet.title == "FileSets":
                        for index in range(len(row_set)):
                            if index == 0:
                                row_set_dict["Index"] = row_set_count
                            elif index == 1:
                                row_set_dict["FileSetName"] = row_set[index]
                            elif index == 2:
                                row_set_dict["Includes"] = row_set[index]
                            elif index == 3:
                                row_set_dict["Excludes"] = row_set[index]
                            elif index == 5:
                                row_set_dict["Recurse"] = row_set[index]
                            else:
                                pass

                        IncludesRead.append(row_set_dict)
                        row_set_count += 1
        return IncludesRead

    def write_files_sizes(self, fs_dict):
        """
        This method writes a supplied dictionary to the FileSizes sheet of the BackupList.xlsx workbook using pandas.
        :param fs_dict:
        :return:
        """
        wb = load_workbook(os.path.join(self.resource_path, "BackupList.xlsx"))
        writer = pd.ExcelWriter(os.path.join(self.resource_path, "BackupList.xlsx"), engine='openpyxl')
        writer.book = wb
        writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)

        ws = wb["FS_SIZE"]
        for row in ws["B2:C1100"]:
            for cell in row:
                cell.value = None
        writer.save()

        df = pd.DataFrame(data=fs_dict)
        df.to_excel(writer, sheet_name="FS_SIZE", startcol=0, startrow=0,
                    columns=['Path', 'Size'],
                    index_label='Index')
        writer.save()
        return df.size
