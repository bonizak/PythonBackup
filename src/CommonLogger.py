import datetime
import logging
import os
import sys
from pathlib import Path

from openpyxl import load_workbook


class LoggerServices:
    """
    This class contains the various methods needed to log and notify
    messages in a log file for each script run
    
     Args
        Required: none
        Optional: none

    Logging: INFO
    
    """

    __author__ = "Barry Onizak"
    __version__ = "20210815.1"
    # # # # # End of header # # # #

    log_file = ""
    log_level = ""

    @staticmethod
    def date():
        """
        This method returns a formatted string of the date in YYYY/MM/DD/hh/mm/ss format
        """
        return datetime.datetime.now().strftime('%Y/%m/%d %H:%M:%S')

    @staticmethod
    def file_date():
        """
        This method returns a formatted date string in YYYYMMDD for appending to file names
        """
        return datetime.datetime.now().strftime('%Y%m%d')

    def setLogFile(self):
        """
        This method sets log file to be written to for each script run
        """
        return os.path.join(self.getLogDir(),
                            f'{str(os.path.basename(sys.argv[0])).replace(".py", "")}.{self.file_date()}.log')

    def getLogDir(self):
        """
        This method sets the logs directory
        """
        log_dir = os.path.join(str(Path.home()), 'logs', self.getScriptName())

        if os.path.isdir(log_dir):
            return log_dir
        else:
            os.makedirs(log_dir)
            if not os.path.exists(log_dir):
                msg = 'Initial setup of logs dir not done, exiting script run!'
                print(f'{self.date()}: {msg}')
                sys.exit(1)
            else:
                return log_dir

    def getLogger(self, name):
        """
        This method creates and returns a object used to log each script run
        """
        log_file = self.openlogfile()
        self.log_level = self.set_log_level()
        logger = logging.getLogger(name)
        logging.basicConfig(filename=log_file, level=self.log_level,
                            format=' %(asctime)s %(levelname)s: %(message)s', datefmt='%Y/%m/%d %H:%M:%S')

        return logger

    def starting_template(self, parameter_list):
        """
        This method takes a list of cmd line args
        passed and displays each running script in a similar view
        """
        self.info(f'Starting {self.getScriptName()}')
        self.info(f"Extracting input params: {(' '.join(map(str, parameter_list)))}")
        self.info(f"Log Level {self.get_log_level()}")
        return None

    def get_log_level(self):
        return self.log_level

    def openlogfile(self):
        """
        This method opens the logfile and prepends the starting info
        """
        self.log_file = self.setLogFile()
        try:
            if os.path.exists(self.log_file):
                fo = open(self.log_file, 'a', encoding='utf-8')
                print(f'\nUsing logfile {self.log_file}')
            else:
                fo = open(self.log_file, 'w', encoding='utf-8')
                print(f'\nCreated logfile {self.log_file}')

            fo.write("\n\n")
            fo.write(self.startScriptLine())
            fo.write("\n")
            fo.close()
        except IOError as ioe:
            msg = f'Log file {self.log_file} write error. {ioe}.'
            print(f'{self.date()}: {msg}')

        return self.log_file

    def closelogfile(self):
        """
        This method appends the closing log info and closes the logfile
        """
        self.log_file = self.setLogFile()
        try:
            fo = open(self.log_file, 'a', encoding='utf-8')
            fo.write("\n")
            fo.write(self.endingScriptLine())
            fo.write("\n")
            fo.close()
        except IOError as ioe:
            msg = f'Log file {self.log_file} write error. {ioe}.'
            print(f'{self.date()}: {msg}')

        return self.log_file

    @staticmethod
    def getScriptName():
        """
        This method returns the script name
        """
        return str(os.path.basename(sys.argv[0])).split('.')[0]

    def startScriptLine(self):
        """
        This method returns the script logfile opening info
        """
        script_path = os.path.normpath(os.path.join(os.popen("pwd").read().strip('\n'), str(sys.argv[0])))
        if not os.path.isfile(script_path):
            print('No such script name in toolkit folder...exiting')
            sys.exit(1)

        script_start = f'{self.separationBar()} \n Starting script {script_path} \n{self.separationBar()}'
        return script_start

    def endingScriptLine(self):
        """
        This method takes a list of cmd line args
        passed and displays each running script in a similar view
        """
        script_path = os.path.normpath(os.path.join(os.popen("pwd").read().strip('\n'), str(sys.argv[0])))
        script_end = f'{self.separationBar()} \n End of script {script_path} \n{self.separationBar()}'
        return script_end

    def set_log_level(self):
        resource_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "resource")
        wb = load_workbook(os.path.join(resource_path, "BackupList.xlsx"))
        sheetset = {'AppConfig': 6}

        for ws in wb:
            if ws.title in sheetset.keys():
                worksheet = wb[ws.title]

                row_sets = [worksheetsets for worksheetsets in worksheet.iter_rows(
                    min_row=1, max_col=sheetset[ws.title], min_col=1, values_only=True)]

                self.log_level = row_sets[3][1]
        return self.log_level

    @staticmethod
    def separationBar():
        """
        This method returns a 30 char separation bar
        """
        return 75 * f'='

    def critical(self, msg):
        """
        This method takes a message and logs it as a
        CRITICAL to the script run log
        """
        logging.critical(msg)
        return None

    def error(self, msg):
        """
        This method takes a message and logs it as an
        ERROR  to the script run log
        """
        logging.error(msg)
        return None

    def warn(self, msg):
        """
        This method takes a message and logs it as an
        WARNING to the script run log
        """
        logging.warning(msg)
        return None

    def info(self, msg):
        """
        This method takes a message and logs it as an
        INFO to the script run log
        """
        logging.info(msg)
        return None

    def debug(self, msg):
        """
        This method takes a message and logs it as a
        DEBUG message to the script run log
        """
        logging.debug(msg)
        return None
