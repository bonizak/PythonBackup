import os
import sys

import pandas as pd
from openpyxl import *
from CommonOs import OsServices as os_services


class ReloadFileSets(os_services):
    """
        This class contains the methods to collect and create a list of all directories, to the level of directories
        as provided, under the supplied list of file systems' roots. It will then write the list into the
        'INCLUDES', 'EXCLUDES , 'COMPRESS' and 'RECURSE columns of
        FILESETS sheet of BACKUPLIST.xls.
    """

    __author__ = "Barry Onizak"
    __version__ = "20220328.2"
    # # # # # End of header # # # #

    resource_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "resource")

    def Build_FileSets(self):
        """
        This method starts the process to read a list of RootPaths and scan them for sub-directories and files,
        then writing these files and folders into the FileSets sheet of the BackupList.xlsx workbook.

        :return: FileSetRows[]
        """
        FileSetRows = []
        FileSystemsIn = self.read_rootpaths_in()
        if len(FileSystemsIn) > 0:
            for fsi in FileSystemsIn:
                for dirpath, subdirList, filesList in self.walklevel(fsi["RootPath"], fsi["MaxDepth"]):
                    if fsi["MaxDepth"] == 0:
                        ''' empty scan '''
                        pass
                    elif fsi["MaxDepth"] == -1:
                        ''' collect the files in the RootFileSystem directory and all subdirectories'''
                        if fsi["FilesFolders"] in ("Folders", "Both"):
                            for subdir in subdirList:
                                if os.path.join(dirpath, subdir).find("Trash") == -1 and \
                                        os.path.join(dirpath, subdir).find("lost+found") == -1 and \
                                        not os.path.islink(os.path.join(dirpath, subdir)):
                                    FileSetRows.append(
                                        {"FileSetName": "ToBeUpdated",
                                         "Includes": os.path.join(dirpath, subdir),
                                         "Excludes": "NA", "Compress": "YES", "Recurse": "Yes"})
                        if fsi["FilesFolders"] in ("Files", "Both"):
                            for file in filesList:
                                if os.path.join(dirpath, file).find("Trash") == -1 and \
                                        os.path.join(dirpath, file).find("lost+found") == -1 and \
                                        not os.path.islink(os.path.join(dirpath, file)):
                                    FileSetRows.append(
                                        {"FileSetName": "ToBeUpdated",
                                         "Includes": os.path.join(dirpath, file),
                                         "Excludes": "NA", "Compress": "YES", "Recurse": "No"})
                    elif fsi["MaxDepth"] == os.path.abspath(fsi["RootPath"]).count(os.path.sep):
                        ''' collect the files and directories ONLY in the RootFileSystem - no recursive scan'''
                        if fsi["FilesFolders"] in ("Folders", "Both"):
                            for subdir in subdirList:
                                if os.path.join(dirpath, subdir).find("Trash") == -1 and \
                                        os.path.join(dirpath, subdir).find("lost+found") == -1 and \
                                        not os.path.islink(os.path.join(dirpath, subdir)):
                                    if os.path.join(dirpath, subdir).count(os.path.sep) <= fsi["MaxDepth"]:
                                        FileSetRows.append(
                                            {"FileSetName": "ToBeUpdated",
                                             "Includes": os.path.join(dirpath, subdir),
                                             "Excludes": "NA", "Compress": "YES", "Recurse": "No"})
                        if fsi["FilesFolders"] in ("Files", "Both"):
                            for file in filesList:
                                if os.path.join(dirpath, file).find("Trash") == -1 and \
                                        os.path.join(dirpath, file).find("lost+found") == -1 and \
                                        not os.path.islink(os.path.join(dirpath, file)):
                                    if os.path.join(dirpath, file).count(os.path.sep) == fsi["MaxDepth"] + 1:
                                        FileSetRows.append(
                                            {"FileSetName": "ToBeUpdated",
                                             "Includes": os.path.join(dirpath, file),
                                             "Excludes": "NA", "Compress": "YES", "Recurse": "No"})
                    elif fsi["MaxDepth"] >= os.path.abspath(fsi["RootPath"]).count(os.path.sep):
                        ''' collect the files and directories in the RootFileSystem - with recursive scan'''
                        if fsi["FilesFolders"] in ("Folders", "Both"):
                            for subdir in subdirList:
                                if os.path.join(dirpath, subdir).find("Trash") == -1 and \
                                        os.path.join(dirpath, subdir).find("lost+found") == -1 and \
                                        not os.path.islink(os.path.join(dirpath, subdir)):
                                    if os.path.join(dirpath, subdir).count(os.path.sep) <= fsi["MaxDepth"]:
                                        FileSetRows.append(
                                            {"FileSetName": "ToBeUpdated",
                                             "Includes": os.path.join(dirpath, subdir),
                                             "Excludes": "NA", "Compress": "YES", "Recurse": "YES"})
                        if fsi["FilesFolders"] in ("Files", "Both"):
                            for file in filesList:
                                if os.path.join(dirpath, file).find("Trash") == -1 and \
                                        os.path.join(dirpath, file).find("lost+found") == -1 and \
                                        not os.path.islink(os.path.join(dirpath, file)):
                                    if os.path.join(dirpath, file).count(os.path.sep) == fsi["MaxDepth"] + 1:
                                        FileSetRows.append(
                                            {"FileSetName": "ToBeUpdated",
                                             "Includes": os.path.join(dirpath, file),
                                             "Excludes": "NA", "Compress": "YES", "Recurse": "No"})

        if len(FileSetRows) == 0:
            os_services.critical(self, f' Empty RootPath Scan Returned. '
                                       f'Verify MaxDepth values in RootPath Sheet of workbook. '
                                       f'\n Exiting!')
            sys.exit(-1)

        sorted_FileSetRows = sorted(FileSetRows, key=lambda I: I["Includes"])
        write_filesets_rc = self.write_filesets(sorted_FileSetRows)
        os_services.info(self, f' \nCaptured {write_filesets_rc} file sets')
        return FileSetRows

    def read_rootpaths_in(self):
        """
        This method reads in the RootPath sheet from BackupList.xlsx workbook in the
        resources directory of this project, and loads the sheet into a class level dictionary.

        :return: Count of rows read in and written to  FileSystemsIn{}
        """
        resource_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "resource")
        wb = load_workbook(os.path.join(resource_path, "BackupList.xlsx"))
        sheetset = {'RootPaths': 4}
        FileSystemsIn = []

        for ws in wb:
            if ws.title in sheetset.keys():
                worksheet = wb[ws.title]

                row_sets = [worksheetsets for worksheetsets in worksheet.iter_rows(
                    min_row=2, max_col=sheetset[ws.title], min_col=1, values_only=True)
                            if None not in worksheetsets]
                row_set_count = 0
                for row_set in row_sets:
                    row_set_dict = {}
                    if worksheet.title == "RootPaths":
                        for index in range(len(row_set)):
                            if index == 0:
                                row_set_dict["Index"] = row_set_count
                            elif index == 1:
                                row_set_dict["RootPath"] = row_set[index]
                            elif index == 2:
                                row_set_dict["MaxDepth"] = row_set[index]
                            elif index == 3:
                                row_set_dict["FilesFolders"] = row_set[index]
                            else:
                                raise AttributeError

                        FileSystemsIn.append(row_set_dict)
                        row_set_count += 1
                    else:
                        os_services.critical(self, f'No RootPaths sheet found')
        return FileSystemsIn

    def write_filesets(self, fs_dict):
        """
        This method writes a supplied dictionary to the FileSets sheet  of the BackupList.xlsx workbook using pandas.
        :param fs_dict:
        :return:
        """
        try:
            wb = load_workbook(os.path.join(self.resource_path, "BackupList.xlsx"))
            writer = pd.ExcelWriter(os.path.join(self.resource_path, "BackupList.xlsx"), engine='openpyxl')
            writer.book = wb
            writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)

            ws = wb["FileSets"]
            for row in ws["C2:F10000"]:
                for cell in row:
                    cell.value = None
            writer.save()

            df = pd.DataFrame(data=fs_dict)
            df.to_excel(writer, sheet_name="FileSets", startcol=1, startrow=0,
                        columns=['FileSetName', 'Includes', 'Excludes', 'Compress', 'Recurse'],
                        index=False)
            writer.save()
        except IOError as ioe:
            os_services.critical(self, f'IO error writing to "BackupList.xlsx : {ioe} ')
        return df.size

    def walklevel(self, rfspath, max_depth):
        """
            This method takes a path and a max_depth value and
            walks the path to max_depth collecting and
            yielding directories and files
            :param rfspath: Root File system to be scanned
            :param max_depth: Maximum directory depth to scan the path
            :return: yield tuples of max_depth, root, dirs[:], files
        """
        if max_depth == 0:  # return nothing
            return
        elif max_depth < 0:  # return everything
            for dirpath, dirnames, filenames in os.walk(rfspath, followlinks=False):
                yield dirpath, dirnames[:], filenames
            return
        else:  # return only up to the max depth
            rfspath = str(rfspath).rstrip(os.path.sep)  # strip of any ending pathsep
            if os.path.isdir(rfspath):
                rfspath_depth = rfspath.count(os.path.sep)
                for dirpath, dirnames, filenames in os.walk(rfspath, followlinks=False):
                    dirpath_depth = dirpath.count(os.path.sep)
                    yield dirpath, dirnames, filenames
                    if rfspath_depth + max_depth <= dirpath_depth:
                        del dirnames[:]
            else:
                os_services.critical(self, f'{rfspath} is not a valid directory')

    def build_file_sizes(self, file_path_dict):
        """
        This method copies all files or folders from FileSets.Includes and GeneralList.SourceFile_FolderName
        into FS_SIZE.PATH and callecting the size of same to write into the Size field.
        :return:
        """
        pass
