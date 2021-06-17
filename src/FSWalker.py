import os
import sys
import pandas as pd

from openpyxl import load_workbook
from CommonOs import OsServices as os_services


class FSWalker(os_services):
    """
        This class contains the methods to collect and create a list of all directories, to the level of directories
        as provided, under the supplied list of file systems roots. It will then write the list into the
        'INCLUDES', 'EXCLUDES , 'COMPRESS' and 'RECURSE columns of
        FILESETS sheet of BACKUPLIST.xls.
    """

    resource_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "resource")
    FileSystemsIn = []
    FileSetRows = []

    def Build_FileSets(self):
        if self.read_fs_in() > 0:
            for fsi in self.FileSystemsIn:
                for cur_depth, dirName, subdirList, fileList in self.walklevel(fsi["RootFileSystems"], fsi["MaxDepth"]):
                    if fsi["MaxDepth"] == -1:
                        ''' collect the files in the RootFileSystem directory and all subdirectories'''
                        self.FileSetRows.append(
                            {"Includes": dirName, "Excludes": "NA", "Compress": "YES", "Recurse": "No"})
                    elif fsi["MaxDepth"] == 0:
                        ''' empty scan '''
                        pass
                    elif fsi["MaxDepth"] == cur_depth and len(fileList) > 0:
                        ''' collect only the files in the RootFileSystem with recursive scan '''
                        for rfs_file in fileList:
                            self.FileSetRows.append(
                                {"Includes": os.path.join(dirName, rfs_file), "Excludes": "NA", "Compress": "YES",
                                 "Recurse": "NO"})
                    elif cur_depth <= fsi["MaxDepth"] and len(fileList) > 0:
                        ''' collect the files in the RootFileSystem directory and all subdirectories'''
                        self.FileSetRows.append(
                            {"Includes": dirName, "Excludes": "NA", "Compress": "YES", "Recurse": "NO"})
                        for sdir in subdirList:
                            sdir_path = os.path.join(dirName, sdir)
                            if os.path.isdir(os.path.abspath(sdir_path)):
                                if sdir_path.rstrip(os.path.sep).count(os.path.sep) <= fsi["MaxDepth"]:
                                    self.FileSetRows.append(
                                        {"Includes": sdir_path, "Excludes": "NA", "Compress": "YES", "Recurse": "YES"})
                    elif cur_depth <= fsi["MaxDepth"] and len(fileList) == 0:
                        ''' collect the subdirectories in the RootFileSystem directory'''
                        for sdir in subdirList:
                            sdir_path = os.path.join(dirName, sdir)
                            if os.path.isdir(os.path.abspath(sdir_path)):
                                if sdir_path.rstrip(os.path.sep).count(os.path.sep) <= fsi["MaxDepth"]:
                                    self.FileSetRows.append(
                                        {"Includes": sdir_path, "Excludes": "NA", "Compress": "YES", "Recurse": "YES"})
                    else:
                        ''' empty scan '''
                        pass

        self.write_fs(self.FileSetRows)
        os_services.info(self, f' \nFSWalker captured {len(self.FileSetRows)} file sets')

    def read_fs_in(self):
        """
        This method reads in the RootFileSystems sheet from BackupList.xlsx workbook in the
        resources directory of this project, and loads the sheet into a class level dictionary.

        :return: Count of rows read in and written to  FileSystemsIn{}
        """
        wb = load_workbook(os.path.join(self.resource_path, "BackupList.xlsx"))
        sheetset = {'RootFileSystems': 3}

        for ws in wb:
            if ws.title in sheetset.keys():
                worksheet = wb[ws.title]

                row_sets = [worksheetsets for worksheetsets in worksheet.iter_rows(
                    min_row=2, max_col=sheetset[ws.title], min_col=1, values_only=True)
                            if None not in worksheetsets]
                row_set_count = 0
                for row_set in row_sets:
                    row_set_dict = {}
                    if worksheet.title == "RootFileSystems":
                        for index in range(len(row_set)):
                            if index == 0:
                                row_set_dict["Index"] = row_set_count
                            elif index == 1:
                                row_set_dict["RootFileSystems"] = row_set[index]
                            elif index == 2:
                                row_set_dict["MaxDepth"] = row_set[index]
                            else:
                                raise AttributeError

                        self.FileSystemsIn.append(row_set_dict)
                        row_set_count += 1
        return row_set_count

    def write_fs(self, fs_dict):
        """
        This method writes a supplied dictionary to the FileSets sheet  of the BackupList.xlsx workbook using pandas.
        :param fs_dict:
        :return:
        """

        df = pd.DataFrame(data=fs_dict)
        df.to_excel('BackupList.xlsx', sheet_name="FileSets", startcol="B")

    def walklevel(self, path, max_depth):
        """
            This method takes a path and a max_depth value and
            walks the path to max_depth collecting and
            yielding directories and files
        :param path: Root File system to be scanned
        :param max_depth: Maximum directory depth to scan the path
        :return: yield tuples of max_depth, root, dirs[:], files
        """
        if max_depth < 0:
            for root, dirs, files in os.walk(path):
                yield max_depth, root, dirs[:], files
            return
        elif max_depth == 0:
            return

        base_depth = path.rstrip(os.path.sep).count(os.path.sep)
        for root, dirs, files in os.walk(path):
            cur_depth = root.count(os.path.sep)
            yield cur_depth, root, dirs[:], files
            # if base_depth + max_depth <= cur_depth:
            if cur_depth <= max_depth:
                del dirs[:]
